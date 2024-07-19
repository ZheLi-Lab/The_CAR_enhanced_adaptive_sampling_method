import numpy as np
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Border
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('agg')
from .segment_related_tools import Segment_lambda_tools
from .plotting import PLOTTING
from .FE_cal_convergence import CAL_FREE_ENERGY


from contextlib import contextmanager
@contextmanager
def bf_af_plot(mkdir_,):
    pwd = os.getcwd()
    if os.path.exists(mkdir_) and os.path.isdir(mkdir_):
        pass
    else:
        os.makedirs(mkdir_)
    os.chdir(mkdir_)
    yield
    os.chdir(pwd) 


class AnalyseSegment():
    def __init__(self, data_df, iffep=True, ifbar=False, scale_f=0.8):
        '''
        Parameters
        ----------
        data_df: pd.DataFrame
            The data contains energy information from one segment lambdas, obtained by ReadOut progress.
        scale_f: float
            To tail the percentage of the every dataframe data used for the calculation. Default: 0.8
        
        Key properties
        ----------
        self.cal_fe_obj: CAL_FREE_ENERGY object
            The CAL_FREE_ENERGY object that contains all free energy related features and functions.
        self.data_df: pd.DataFrame
            The dataframe obtained by concating all data values
        '''
        self.cal_fe_obj = CAL_FREE_ENERGY(data_df, False, scale_f, True)
        self.data_df = self.cal_fe_obj.all_data_unk
        if iffep:
            self.cal_fe_and_reweight()
        if ifbar:
            self.cal_fe_BAR()

    def cal_fe_BAR(self,):
        self.bar_fe_df = self.cal_fe_obj.cal_FE(None, unit='kcal/mol')

    def cal_fe_and_reweight(self):
        '''
        Calculate free energy and do reweighting to get reweighed energy information.
        
        Key properties
        ----------
        self.reweight_fe_df: pd.DataFrame
            indexs: free_energy_what_to_what
            columns: witch nearby window used to calculate free energy. (win_n, where n is lambda{use_dU_to_reweight}-lambda{target_for_reweighting_to})
        self.dG_diff_df: pd.DataFrame
            The dataframe in format of diagonal sheet, obtained by substraction of reweighted_free_energy and original_free_energy.
        self.dG_diffPercent_df: pd.DataFrame
            The dataframe in format of diagonal sheet, obtained by (reweighted_free_energy - original_free_energy)/abs(original_free_energy).
        '''
        ## fep forward
        self.reweight_forward_df, self.dG_forward_diff_df, self.dG_forward_diffPercent_df = self.cal_fe_obj.cal_reweight_all(use_wins_step=0, use_wins_num=3, unit='kcal/mol', ifcal_sum=True, ifdiagnum=True)
        ## fep reverse
        self.reweight_reverse_df, self.dG_reverse_diff_df, self.dG_reverse_diffPercent_df = self.cal_fe_obj.cal_reweight_all(use_wins_step=0, use_wins_num=3, unit='kcal/mol', forward=False, ifcal_sum=True, ifdiagnum=True)

    def plot_reweight_heatmap(self, error_max_edge, png_file_postfix, ifforward=True, ifPercent=False, ifplot=True, iftocsv=False):
        '''
        Parameters
        ----------
        error_max: float
            Specify the tolerance free energy error of a single thermodynamic process.
            To determine the boundary values of color scaleplate in reweighting heatmap.
        png_file_postfix: str
            The postfix of the output heatmap pngfile' name.
            Output file will be {first_win_lambda}_to_{last_win_lambda}_heatmap_diff(or diffPercent)_{png_file_postfix}.png
        '''
        if ifPercent:
            if ifforward:
                df = self.dG_forward_diffPercent_df
                heatmap_form = 'forward_diffPercent'
            else:
                df = self.dG_reverse_diffPercent_df
                heatmap_form = 'reverse_diffPercent'
        else:
            if ifforward:
                df = self.dG_forward_diff_df
                heatmap_form = 'forward_diff'
            else:
                df = self.dG_reverse_diff_df
                heatmap_form = 'reverse_diff'
        start_coords, end_coords = Segment_lambda_tools.get_lambda_start_end_coords(self.reweight_forward_df.index[-1])
        delta_lambda, label, lambda_info_x = Segment_lambda_tools.cal_delta_lambda(start_coords, end_coords)
        if ifplot:
            error_max = np.around(error_max_edge*(abs(delta_lambda)**(1/2)), decimals=3)
            plt.clf()
            plot_obj = PLOTTING()
            plot_obj.plot_heatmap_cmap(df, error_max, f'{start_coords}_to_{end_coords}_{heatmap_form}_{png_file_postfix}.png')
        if iftocsv:
            with bf_af_plot('reweight_csv'):
                df.to_csv(f'{start_coords}_to_{end_coords}_{heatmap_form}_{png_file_postfix}.csv', sep='|')
        plt.close()

class SegmentConvergenceAnalysis():
    def __init__(self, init_segment_lambda, frames_each_simu, error_max_edge=0.2, ifrun_turnaround=True, ifoverwrite=False, prefix=''):
        '''
        Parameters
        ----------
        init_segment_lambda: dict
            The segment lambdas in initial order that refers to forward simulation direction.
        frames_each_simu: int

        error_max_edge: folat, in unit of kcal/mol.
            Tolerance error in predicting the free energy of a thermodynamic process (lambda 0 to 1).
         
        Key properties
        ----------
        self.init_segment_lambda: dict
            The segment lambdas in initial order that refers to forward simulation direction.
        self.forward_segment_lambda: dict
            The segment lambdas in initial order that refers to forward simulation direction.
        self.reverse_segment_lambda: dict
            The segment lambdas in reverse order that refers to backward simulation direction.
        self.frames_each_simulation: int

        self.count: int
            Repeated times of the segment simulation. Each direction contributes one time.
            e.g. 0.0 -> 0.1 (count=0) and 0.1 -> 0.0 (count=1)
        self.error_max_edge: float, in unit of kcal/mol.
            Tolerance error in predicting the free energy of a thermodynamic process (lambda 0.0 to 1.0).
        self.error_segment: float, error_segment = error_max_edge*(delta_lambda**(1/2))
            Tolerance error in predicting the free energy of the current segment (part lambda).
        self.fe_eachwin_all: dict
            The dict to store free energy results dataframe after each run in detail (containing every window's result).
        self.compare_time_serials_all: 

        self.changes_result_df: pd.DataFrame
            The dataframe to store analysis results , recording the convergency related feature with simulation repeating.
        '''
        ### initialize simulation lambda dict for later md running and count simulation repetition times.
        self.init_segment_lambda = init_segment_lambda
        self.forward_segment_lambda, self.reverse_segment_lambda = self.get_simu_lambda(init_segment_lambda, ifrun_turnaround)
        self.frames_each_simulation = frames_each_simu
        ### calculate the tolerance
        self.error_max_edge = error_max_edge
        self.start_coords = [init_segment_lambda["lambda_restraints"][0], init_segment_lambda["lambda_electrostatics"][0], init_segment_lambda["lambda_sterics"][0]]
        self.end_coords = [init_segment_lambda["lambda_restraints"][-1], init_segment_lambda["lambda_electrostatics"][-1], init_segment_lambda["lambda_sterics"][-1]]
        self.lambda_what_to_what = f"{self.start_coords[0]}_{self.start_coords[1]}_{self.start_coords[2]}_to_{self.end_coords[0]}_{self.end_coords[1]}_{self.end_coords[2]}"
        delta_lambda, label, lambda_info_x = Segment_lambda_tools.cal_delta_lambda(self.start_coords, self.end_coords)
        print(f'start_cor:{self.start_coords}, end_cor:{self.end_coords}, delta lambda is {delta_lambda}')
        self.error_segment = error_max_edge*(abs(delta_lambda))
        self.prefix = prefix

        ### result file
        with bf_af_plot(f'{self.prefix}segment_converge_analysis'):
            if ifoverwrite or not os.path.exists(f"{self.prefix}{self.lambda_what_to_what}_analyse_convergence.xlsx"):
                self.count = 0
                self.count_0 = 0
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'analyse_convergence'
                wb.save(f"{self.prefix}{self.lambda_what_to_what}_analyse_convergence.xlsx")
            else:
                changes_result_df = pd.read_excel(f"{self.prefix}{self.lambda_what_to_what}_analyse_convergence.xlsx", sheet_name='analyse_convergence', index_col=[0])
                self.count = len(changes_result_df)
                self.count_0 = len(changes_result_df)

        
    def get_simu_lambda(self, init_segment_lambda, ifrun_turnaround=True):
        '''
        Get forward simulation lambdas dict and reverse simulation lambdas dict.

        Parameters
        ----------
        init_segment_lambda: dict
            The segment lambdas in initial order that refers to forward simulation direction.
        ifrun_turnaround: bool
            
        
        Return
        ----------
        forward_segment_lambda: dict
            The segment lambdas in initial order that refers to forward simulation direction.
        reverse_segment_lambda: dict
            The segment lambdas in reverse order that refers to backward simulation direction.
        '''
        forward_segment_lambda = {}
        reverse_segment_lambda = {}
        if ifrun_turnaround:
            for key, value in init_segment_lambda.items():
                forward_segment_lambda[key] = value[:]
                reverse_segment_lambda[key] = value[::-1]
        else:
            for key, value in init_segment_lambda.items():
                reverse_segment_lambda[key] = value[-2::-1]
                forward_segment_lambda[key] = value[1::]
        return forward_segment_lambda, reverse_segment_lambda

    def updata_xlsx(self, writer, single_df, sheetname):
        single_df = single_df.rename_axis('count')
        try:
            pre_df = pd.read_excel(f"{self.prefix}{self.lambda_what_to_what}_analyse_convergence.xlsx", sheet_name=sheetname, index_col=[0])
            df_new = pd.concat([pre_df, single_df])
            df_new.to_excel(writer, sheet_name=sheetname)
        except:
            single_df.to_excel(writer, sheet_name=sheetname)

    def updata_fe_properties(self, ana_type, ana_objs, init_fe, writer):
        fe_data = {}
        for i in range(1, ana_objs.num+1):
            ana_obj = getattr(ana_objs, f'ana_obj_{ana_type}{i}')
            single_df = pd.DataFrame([list(ana_obj.bar_fe_df.iloc[:,0])], index=[self.count], columns=ana_obj.bar_fe_df.index)
            single_df = single_df.rename_axis('count')
            self.updata_xlsx(writer, single_df, f'fe_data_{ana_type}{i}')
            fe_data[f'fe_{ana_type}{i}'] = np.around(ana_obj.bar_fe_df.iloc[-1,0], decimals=3)
        fe_data_arr = list(fe_data.values())
        fe_data[f'moving_estimator_stability'] = np.around(np.std([init_fe]+list(fe_data_arr)), decimals=4)
        # for i in range(len(fe_data_arr)-1):
        #     fe_data[f'fe_diff_{ana_type}_{i+2}-{i+1}'] = fe_data_arr[i+1] - fe_data_arr[i]
        # fe_data[f'fe_diff_RMS_{ana_type}'] = np.around(np.sqrt(np.mean(np.array(list(fe_data.values())[-len(fe_data_arr)+1:])**2)), decimals=4)
        return fe_data

    def compare_latest_moving_parts_percent(self, data_dict, init_fe, writer, part_num=4):
        '''
        Compare free energy analyzed with different parts of data (60%~80% and 80%~100%).
        The maximum number of each part's frames to be analysed is the number of 500 ps md production frames.
        '''
        ana_type = 'moving'
        class ANA:
            num = part_num
        ana_objs = ANA()
        if part_num <= 4:
            width_ratio = 0.1
        else:
            width_ratio = 0.4/part_num
            raise Warning(f"The number of moving slides to be analysed is {part_num}, and the width ratio is set to 0.8/part_num.")
        for i in range(1,part_num+1):
            # print(f'-----###### need to get {part_num} parts, now getting part {i} ######-----')
            df_list = []
            for lambda_, single_df in data_dict.items():
                len_df = len(single_df)
                start_frame = int(np.floor(len_df - (part_num+1-i)*width_ratio*len_df))
                end_frame = int(np.floor(len_df - (part_num-i)*width_ratio*len_df))
                df_list.append(single_df.iloc[start_frame:end_frame, :])
                # print(f'lambda:{lambda_}, original_df_len:{len_df}\nstart_frame:{start_frame}, end_frame:{end_frame}, ana_df_len:{len(df_list[-1])}')
            df = pd.concat(df_list)
            setattr(ana_objs, f"ana_obj_{ana_type}{i}", AnalyseSegment(df, False, True, scale_f=1))
        fe_result = self.updata_fe_properties(ana_type, ana_objs, init_fe, writer)
        compare_result = fe_result.copy()
        return compare_result

    def compare_forward_reverse_time_serials(self, data_dict, writer, part_num=10, compare_scale=0.8):
        '''
        Compare the free energy results of forward and reverse time serials after time_ratio=0.5.
        '''
        class ANA:
            pass
        ana_objs = ANA()
        fe_diff_all = []
        score = 0
        cur_ratio = 0.5
        divided_ratio = (compare_scale-cur_ratio)/part_num
        while (cur_ratio <= compare_scale):
            for direct in ['forward', 'reverse']:
                setattr(ana_objs, f'df_{direct}_list', [])
            # print(f'--------current ratio: {cur_ratio}----------')
            for lambda_, single_df in data_dict.items():
                len_df = len(single_df)
                forward_end_frame = int(np.floor(len_df*cur_ratio))
                reverse_start_frame = int(np.floor(len_df*(1-cur_ratio)))
                df_f = single_df.iloc[0:forward_end_frame, :]
                df_b = single_df.iloc[reverse_start_frame:, :]
                # print(f'lambda:{lambda_}, original_df_len:{len_df}\nforward_start_frame:{forward_start_frame}, forward_end_frame:{forward_end_frame}, reverse_start_frame:{reverse_start_frame}\ndf_forward_len:{len(df_f)}, df_reverse_len:{len(df_b)}')
                ana_objs.df_forward_list.append(df_f)
                ana_objs.df_reverse_list.append(df_b)
            # print('________________________________________')
            for direct in ['forward', 'reverse']:
                df_list = getattr(ana_objs, f'df_{direct}_list')
                df = pd.concat(df_list, sort=False)
                cal_fe_obj = CAL_FREE_ENERGY(df, False, 1, True)
                setattr(ana_objs, f'fe_{direct}', cal_fe_obj.cal_FE(None, 'kcal/mol').iloc[-1,0])
            fe_diff = np.around(ana_objs.fe_forward-ana_objs.fe_reverse, decimals=4)
            fe_diff_all.append(fe_diff)
            score = score + int(abs(fe_diff)<=self.error_segment)
            cur_ratio = round(cur_ratio+divided_ratio, 3)
        single_df = pd.DataFrame([fe_diff_all], columns=np.around(np.arange(0.5, compare_scale+0.00001, divided_ratio), decimals=3), index=[self.count])
        single_df = single_df.rename_axis('count')
        self.updata_xlsx(writer, single_df, 'compare_time_serials')
        diff_mean = np.around(np.mean(abs(np.array(fe_diff_all))), decimals=4)
        low_gap_rate = np.around(score/len(fe_diff_all), decimals=4)
        return {'fe_diff_ratio50':fe_diff_all[0],'forward-reverse_estimator_coincidence': diff_mean, 'time_serials_low_gap_rate':low_gap_rate}

    def cal_ene_diff(self, fe, change_result_df, fe_col_surfix):
        try:
            prev_fe_result_list = list(change_result_df[f'fe_{fe_col_surfix}'])
            fe_diff = np.around(fe - prev_fe_result_list[-1], decimals=4)
            last_5_fe_std = np.around(np.std([fe]+list(prev_fe_result_list[-4:])),decimals=4)
        except:
            fe_diff, last_5_fe_std= np.NaN, np.NaN
        return {f'fe_diff_{fe_col_surfix}':fe_diff, f'fe_5std_{fe_col_surfix}':last_5_fe_std}

    def compare_forward_ene(self, data_dict, change_result_df, writer):
        compare_forward_ene = {}
        ana_used_data = pd.concat(list(data_dict.values()))
        cal_fe_obj = CAL_FREE_ENERGY(ana_used_data, False, 1, True)
        fe_data = cal_fe_obj.cal_FE(None, 'kcal/mol')
        fe_data_df = pd.DataFrame([list(fe_data.iloc[:,0])], columns=fe_data.index, index=[self.count])
        self.updata_xlsx(writer, fe_data_df, f'fe_forward')
        fe = np.around(fe_data.iloc[-1,0], decimals=3)
        compare_forward_ene.update({f'fe_forward':fe})
        fe_change_result = self.cal_ene_diff(fe, change_result_df, f'forward')
        compare_forward_ene.update(fe_change_result)
        return compare_forward_ene

    def cal_converge_score(self, data_dict, compare_simu_nums=3, time_serials_num=10):
        '''
        Update convergence metrics and give a score.

        Parameter
        ----------
        
        '''
        with bf_af_plot(f'{self.prefix}segment_converge_analysis'):
            with pd.ExcelWriter(f"{self.prefix}{self.lambda_what_to_what}_analyse_convergence.xlsx", mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                changes_result_df = pd.read_excel(f"{self.prefix}{self.lambda_what_to_what}_analyse_convergence.xlsx", sheet_name='analyse_convergence', index_col=[0])
                ana_result = {}
                compare_forward_result = self.compare_forward_ene(data_dict, changes_result_df, writer)
                ana_result.update(compare_forward_result)
                compare_latest_moving_parts_result = self.compare_latest_moving_parts_percent(data_dict, ana_result[f'fe_forward'], writer, compare_simu_nums)
                ana_result.update(compare_latest_moving_parts_result) 
                compare_two_direct_result = self.compare_forward_reverse_time_serials(data_dict, writer, time_serials_num)
                ana_result.update(compare_two_direct_result)
                fe_forward_converge = int(ana_result[f'fe_5std_forward']<=self.error_segment)
                fe_moving_converge = int(ana_result['moving_estimator_stability']<=self.error_segment*2)
                time_serials_converge = int(ana_result['forward-reverse_estimator_coincidence']<=self.error_segment)
                time_serials_gap_converge = int(ana_result['time_serials_low_gap_rate']>=0.75)
                converge_score = (fe_moving_converge + time_serials_gap_converge) * fe_forward_converge * time_serials_converge
                ana_result['converge_score'] = converge_score
                print(f'----metrics----\n error_segment:{self.error_segment}')
                print(f'converge_score:{converge_score}')
                print(f'fe_std_convergence: forward:{fe_forward_converge}, moving:{fe_moving_converge}; time_serials_converge: diff_mean {time_serials_converge}, low_gap_rate {time_serials_gap_converge}\n------------')
                single_result_df = pd.DataFrame(ana_result, index=[self.count])
                single_result_df = single_result_df.rename_axis('count')
                changes_result_df_new = pd.concat([changes_result_df, single_result_df])
                changes_result_df_new.to_csv(f"{self.prefix}{self.lambda_what_to_what}_analyse_convergence.csv")
                changes_result_df_new.to_excel(writer, sheet_name='analyse_convergence')
        return converge_score

    def convergence_aly(self, data_dict, segment_lambda, min_reitera_times, max_reitera_times, compare_simu_nums=3, time_serials_num=10):
        '''
        min_reitera_times: int
            The minimum number of repeated round-trip simulations, determining the least repetition of the current segment.
            A reverse simulation and a forward simulation form one reiteration.
        max_reitera_times: int
            The maximum number of repeated round-trip simulations, determining when to terminate the repetition and start the next segment.
            A reverse simulation and a forward simulation form one reiteration.
        '''
        print(f'now analyze No.{self.count}')
        ifnext_segment = False
        if self.count-self.count_0 < max_reitera_times*2:
            if segment_lambda == self.forward_segment_lambda or segment_lambda == self.init_segment_lambda:
                segment_lambda = self.reverse_segment_lambda
                converge_score = self.cal_converge_score(data_dict, compare_simu_nums, time_serials_num)
                if self.count >= min_reitera_times*2 and converge_score >= 1:
                    ifnext_segment = True
                    self.output_final_compare_data()
                    self.output_ana_used_data_to_csv(data_dict)
            elif segment_lambda == self.reverse_segment_lambda:
                segment_lambda = self.forward_segment_lambda
                converge_score = self.cal_converge_score(data_dict, compare_simu_nums, time_serials_num)
            else:
                raise ValueError("The segment lambda dict is neither init segment lambda dict nor reverse segment lambda dict.")
        elif self.count-self.count_0 == max_reitera_times*2:
            segment_lambda = self.forward_segment_lambda
            converge_score = self.cal_converge_score(data_dict, compare_simu_nums, time_serials_num)
            ifnext_segment = True
            self.output_final_compare_data()
            self.output_ana_used_data_to_csv(data_dict)
        else:
            raise ValueError("Error: The count is larger than specified maximum reiteration times.")
        self.count = self.count + 1
        return segment_lambda, ifnext_segment

    def output_final_compare_data(self,):
        with bf_af_plot(f'{self.prefix}segment_converge_analysis'):
            with pd.ExcelWriter(f"{self.prefix}{self.lambda_what_to_what}_analyse_convergence.xlsx", mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                changes_result_df = pd.read_excel(f"{self.prefix}{self.lambda_what_to_what}_analyse_convergence.xlsx", sheet_name='analyse_convergence', index_col=[0])
                changes_result_df = changes_result_df.style.highlight_between(color='#D9EAD3',left=[0, 0, -self.error_segment, 0, 0.75],right=[self.error_segment, self.error_segment*2, self.error_segment, self.error_segment, 1.0],
                                                                                subset=[f'fe_5std_forward','moving_estimator_stability','fe_diff_ratio50','forward-reverse_estimator_coincidence','time_serials_low_gap_rate'],axis=1)
                changes_result_df.to_excel(writer, sheet_name='analyse_convergence')

            ### set border to none
            workbook = openpyxl.load_workbook(f"{self.prefix}{self.lambda_what_to_what}_analyse_convergence.xlsx")
            border = Border(top=None, bottom=None, left=None, right=None)
            # workbook = openpyxl.load_workbook('your_file.xlsx')
            for worksheet in workbook.sheetnames:
                ws = workbook[worksheet]
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = border
            workbook.save(f"{self.prefix}{self.lambda_what_to_what}_analyse_convergence.xlsx")
            workbook.close()

    def output_ana_used_data_to_csv(self, df_dict):
        with bf_af_plot("ana_used_data"):
            for lambda_, single_df in df_dict.items():
                single_df.to_csv(f'{self.prefix}lambda{lambda_}.csv', sep='|')